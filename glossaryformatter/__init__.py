import sys
import codecs
import openpyxl


def set_output_encoding(encoding='utf-8'):
    """When piping to the terminal, python knows the encoding needed, and
    sets it automatically. But when piping to another program (for example,
    | less), python can not check the output encoding. In that case, it 
    is None. What I am doing here is to catch this situation for both 
    stdout and stderr and force the encoding.
       
    Thanks to delavnog for providing the function:
    https://stackoverflow.com/questions/19696652/
    """

    current = sys.stdout.encoding
    if current is None :
        sys.stdout = codecs.getwriter(encoding)(sys.stdout)
    current = sys.stderr.encoding
    if current is None :
        sys.stderr = codecs.getwriter(encoding)(sys.stderr)


def get_terms_excel(filename, last_col, sheet='Sheet1', term_col='Term'):
    """Returns a list containing a dict for each row of an Excel spreadsheet.
    The column headers from the spreadsheet are used as the dict keys.

    filename (string): file name of Excel spreadsheet
    sheet (string): sheet name
    term_col (string): header of column containing terms
    """

    wb = openpyxl.load_workbook(filename, read_only=True)
    ws = wb[sheet]

    # put the column headers in a list
    header = []
    for col in range(1, last_col+1):
        header.append(ws.cell(row=1, column=col).value)

    # create a list containing a dict for each term
    terms = []
    for row in ws.iter_rows(min_row=2, max_col=last_col):
        this_term = {}
        n = 0
        for cell in row:
            # use the column header as the dict key
            this_term[header[n]] = cell.value
            n += 1
        # stop if we reach a blank cell where there should be a term
        if not this_term['Term']:
            break
        terms.append(this_term)

    return terms


def get_refs(term, cols = [], indicators = ['Y']):
    """Returns a list of numbers specifying which of the supplied columns
    contains one of the indicators i.e. for which references the term is
    relevant. Reference columns can be used to refer to a particular piece of
    music or chapter in a textbook, for example.
    
    term (dict): a single row which includes the reference columns
    cols (list of strings): column headers for the reference columns
    indicators (list): the possible options for text that indicates a reference
    """

    refs = []
    for n in range(0, len(cols)):
        if term[cols[n]] in indicators:
            refs.append(n + 1)
    return refs


def print_glossary(terms, term_col='Term', def_col='Definition',
                   cat_col='Category', subcat_col=None,
                   omit_col=None, omit_indicator=None,
                   cat_prefix='', cat_suffix='\n',
                   subcat_prefix='', subcat_suffix='\n',
                   empty_subcat_name='Other',
                   begin_terms='', end_terms='',
                   term_prefix='', term_suffix='',
                   def_prefix=' - ', def_suffix='',
                   ref_cols=[], ref_indicators = ['Y'],
                   refs_prefix='', refs_separator=',', refs_suffix=''):
    """Prints a formatted glossary of the supplied terms to stdout. This can
    then be piped to another program or redirected to a file as necessary. The
    default options print a glossary with category names and minimal formatting
    but no sub-category names.
    
    terms (list of dicts): terms obtained using get_terms_excel
    term_col (string): header of column containing terms
    def_col (string): header of column containing definitions
    cat_col (string): header of column containing categories
    subcat_col (string): header of column containing sub-categories
    (sub)cat_prefix / (sub)cat_suffix (string):
        printed before/after (sub-)category names
    empty_subcat_name (string): sub-category name for terms with no sub-category
    begin_terms / end_terms: printed before/after terms in each (sub-)category
    term_prefix / term_suffix (string): printed before/after each term
    def_prefix / def_suffix (string): printed before/after each definition
    ref_cols (list of strings): headers of columns which indicate a reference
        e.g. to a particular piece of music or chapter of a textbook
    ref_indicators (list): possibilities for text that indicates a reference
    refs_prefix / refs_suffix (string): printed before/after reference numbers
    refs_suffix (string): printed in between reference numbers
    """
    
    previous_cat = None
    previous_subcat = None
    no_refs_to_print = False
    term_num = 1

    for this_term in terms:
        # if we're omitting terms and this term has been marked to omit
        if omit_col and this_term[omit_col] == omit_indicator:
            # skip this term and go on to the next
            continue

        # if new category
        if this_term[cat_col] != previous_cat:
            if term_num != 1:
                print(end_terms)
            print(cat_prefix + this_term[cat_col].strip() + cat_suffix)
            # if subcategories column supplied, print it (unless blank)
            if subcat_col:
                if this_term[subcat_col]:
                    print((subcat_prefix + this_term[subcat_col].strip() +
                           subcat_suffix))
            if begin_terms:
                print(begin_terms)

        # if same category as previous term
        else:
            # if subcategories column supplied and subcategory is new
            if subcat_col and this_term[subcat_col] != previous_subcat:
                if term_num != 1:
                    print(end_terms)
                # if subcategory isn't empty
                if this_term[subcat_col]:
                    print((subcat_prefix + this_term[subcat_col].strip() +
                           subcat_suffix))
                else:
                    print((subcat_prefix + empty_subcat_name +
                           subcat_suffix))
                if begin_terms:
                    print(begin_terms)

        # if reference columns supplied
        if len(ref_cols):
            refs = get_refs(this_term, cols = ref_cols,
                            indicators = ref_indicators)
            # if there are some references for this term
            if refs:
                # get string of references separated by separator
                ref_str = refs_separator.join(
                    '{}'.format(num) for num in refs)
                print((term_prefix + this_term[term_col].strip() + term_suffix +
                       def_prefix + this_term[def_col].strip() + def_suffix +
                       refs_prefix + ref_str + refs_suffix))
            else:
                no_refs_to_print = True
        # if no reference columns supplied or they're empty for this term
        if not len(ref_cols) or no_refs_to_print:
            print((term_prefix + this_term[term_col].strip() + term_suffix +
                   def_prefix + this_term[def_col].strip() + def_suffix))

        previous_cat = this_term[cat_col]
        if subcat_col:
            previous_subcat = this_term[subcat_col]
        no_refs_to_print = False
        term_num += 1

    print(end_terms)
